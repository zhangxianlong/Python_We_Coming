#encoding=utf-8

import openpyxl
import os
import datetime
"""
掌握内容
1、了解openpyxl的常用方法，及操作excel的步骤
"""
#创建文件-创建文件对象
wb1 = openpyxl.Workbook()
#保存excel文件
wb1.save(filename = 'excel_name.xlsx')
#打开已有文件
wb2 = openpyxl.load_workbook('excel_name.xlsx')
#激活worksheet
sheet01 = wb2.active
#存储数据方式一：数据可以直接分配到单元格中
sheet01['A2'] = 100
#存储数据方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
sheet01.append([1, 2, 3])
#存储数据方式三：Python 类型会被自动转换
sheet01['A4'] = datetime.datetime.now().strftime("%Y-%m-%d")
#创建表：方式一：插入到最后(default)
sheet02 = wb2.create_sheet('Mysheet')
#创建表：方式二：插入到最开始的位置
sheet03 = wb2.create_sheet('Yoursheet',0)
#选择表：sheet 名称可以作为 key 进行索引
ws1 = wb2['Mysheet']
#选择表：已过时，不建议使用
ws2 = wb2.get_sheet_by_name('Mysheet')
if ws1 is ws2:
    print (True)
else:
    print (False)
#查看表名:直接查找
print (wb2.sheetnames)
#查看表名：遍历查找
for sheet in wb2:
    print (sheet.title)
#访问单元格：单一单元格访问
c = sheet02['A4']
print (c,'这是A4单元格')
#访问单元格：单一单元格访问
d = sheet02.cell(row=4,column=2,value='定位访问单元格')
print (d)
#访问单元格：单一单元格访问
for i in range(1,101):
    for j in range(1,101):
        sheet02.cell(row=i,column=j)
#保存excel
wb2.save(filename='excel_name.xlsx')

#题目：在当前文件夹新建一个excel文件，然后写入数据，读取数据，操作单元格
def Creat_excel(excel_name):
    pwd = os.getcwd()
    print (pwd)
    wb = openpyxl.Workbook()                #创建文件对象
    wb.save(filename = excel_name)          #保存文件，名字为excel01
    print (excel_name + "新建成功")

def Write(data,fields,sheetname,excel_name):
    print ('写入excel...')
    wb = openpyxl.load_workbook(filename = excel_name)      #打开已有的excel文件

    sheet01 = wb.active         #获取第一个sheet
    sheet01.title = sheetname   #设定一个sheet的名字

    field = 1
    for field in range(1, len(fields) + 1):    #写入首行
        row = sheet01.cell(row=1, column=field, value=str(fields[field - 1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入首列
        for col1 in range(1,len(data[row1-2])+1):
            column=sheet01.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=excel_name)
    print("保存成功")

Creat_excel('excel01.xlsx')
Write('asdfgh','qwerty','sheet01','excel01.xlsx')

